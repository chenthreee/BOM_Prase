import os

import openpyxl
import pandas as pd
import requests
import hashlib
import time

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from paserCollection.ParserFactory import ParserFactory

IC_list = ['运算放大器', '微控制器', 'MCU','可编程','处理器','存储器','传感器','放大器']


def refresh_excel(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    rows_to_keep = []
    for row in sheet.iter_rows(min_row=1, max_row=4, values_only=True):
        rows_to_keep.append(row)

    max_row = sheet.max_row
    for row in range(5, max_row + 1):
        sheet.delete_rows(5)

    for i, row_data in enumerate(rows_to_keep, start=1):
        for j, value in enumerate(row_data, start=1):
            sheet.cell(row=i, column=j, value=value)

    workbook.save(excel_path)


def extract_parameters(component_type, description):
    parser = ParserFactory.get_parser(component_type)
    return parser.parse(description)


def get_first_word(s):
    words = s.split()
    return words[0] if words else ''


def url_encode(input_string):
    safe_chars = '-_.~=&|'
    encoded = []

    for c in input_string:
        if c.isalnum() or c in safe_chars:
            # 如果字符是字母、数字或在URL编码中保留的字符，则不进行转码
            encoded.append(c)
        else:
            # 否则，将字符转换为URL编码形式
            encoded.append('%{:02X}'.format(ord(c)))

    return ''.join(encoded)


def md5(value):
    return hashlib.md5(value.encode()).hexdigest()


def token_request(url):
    headers = {
        "Content-Type": "application/json"
    }
    data = {
        "_t": 0,
        "appid": "S$SBQCDZYXGS_565",
        "sign": ""
    }
    timestamp = int(time.time())
    data["_t"] = timestamp
    ready_to_md5 = f"_t={timestamp}&appid=S$SBQCDZYXGS_565|7b40ee9ed472baa1bc1a02a74827217a"
    ready_to_md5 = url_encode(ready_to_md5)
    sign = md5(ready_to_md5)
    data["sign"] = sign
    try:
        result = requests.post(url, headers=headers, json=data, verify=False)
        result.raise_for_status()  # 如果状态码不是200，引发HTTPError异常
        return result.json()
    except requests.ConnectionError:
        print("Error: Failed to connect to the server.")
    except requests.Timeout:
        print("Error: The request timed out.")
    except requests.RequestException as e:
        print(f"Error: An error occurred: {e}")
    except ValueError:
        print("Error: Failed to parse JSON response.")
    return None


def single_query(url, token, part_code):
    temp = part_code
    current_time = int(time.time())
    readytoMD5 = f"_t={current_time}"
    temp = temp.replace(" ", "")  # Remove spaces
    readytoMD5 += f"&keyword={temp}"
    readytoMD5 += f"&token={token}"
    readytoMD5 += "|7b40ee9ed472baa1bc1a02a74827217a"

    readytoMD5 = url_encode(readytoMD5)
    sign = md5(readytoMD5)
    params = {
        '_t': current_time,
        'keyword': temp,
        'sign': sign,
        'token': token
    }
    print(params)
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
    }

    response = requests.post(url, data=params, headers=headers)
    if response.status_code != 200:
        return None

    try:
        # 解析 JSON 响应
        result = response.json().get("result", [])
    except requests.exceptions.JSONDecodeError as e:
        result = None
    # return response.json().get("result", [])
    return result


def token_access():
    token_api_url = "https://openapi.ickey.cn/v2/new-token/create"
    response = token_request(token_api_url)
    if response is None:
        print("Error: Failed to get token.")
        exit(1)
    else:
        return response["result"]['token']


def color_mark(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    light_orange_fill = PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid")  # 更淡的橙色
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 淡灰色
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        cell_2 = row[2].value
        cell_12 = row[12].value

        if cell_12 is None and cell_2 is None:
            for cell in row:
                cell.fill = light_orange_fill
        elif cell_12 is None and cell_2 is not None:
            for cell in row:
                cell.fill = grey_fill
    # 保存修改后的Excel文件
    wb.save(file_path)  # 你可以选择新的文件名保存修改后的文件


def query_res_process(query_res, query_part_code, manufacturer):
    #######On pending ： 可能有返回的有效数据，但是数据某些字段可能为空，需做例外处理
    query_part_code = query_part_code.strip()
    # manufacturer = get_first_word(manufacturer)  # 需要优先选同样厂商的？如果不是同样厂商，参数是否会有所不同
    dip_keywords = ["径向", "通孔", "dip", "插件"]
    smt_keywords = ["贴片", "SMD", "SMT", "贴装"]

    mount_type = ''
    part_type = ''
    pro_desc = ''
    footprint = ''
    datasheet_link = ''
    parameters = {'specification': ''}
    for item in query_res:  # 遍历所有返回结果
        if item.get("pro_name") != query_part_code:   # 如果没有精确查询 是否可行?
            continue
        else:

            type_footprint_desc = (item.get("cate_name") + item.get("footprint") + item.get("pro_desc")).lower()
            if any(keyword in type_footprint_desc for keyword in dip_keywords):
                mount_type = '插件'
            elif any(keyword in type_footprint_desc for keyword in smt_keywords):
                mount_type = '贴片'

            part_type = item.get("cate_name")
            footprint = item.get("footprint")
            pro_desc = item.get("pro_desc")
            datasheet_link = item.get("data_sheet")

            # On Pending ，目前只是处理了第一个料号匹配的查询结果，要达到最好，必须都遍历一遍？
            break

    # 元件类型判断：根据类型以及规格描述中是否包含了某些关键词，执行对应的规格解析函数，待完善
    if "压敏电阻" in part_type or "压敏电阻" in pro_desc:
        parameters = extract_parameters("Varistor", pro_desc)
    elif "安规电容" in part_type or "安规电容" in pro_desc:
        parameters = extract_parameters("Safety Capacitor", pro_desc)
    elif "电阻" in part_type or "电阻" in pro_desc:
        parameters = extract_parameters("Resistor", pro_desc)
    elif "电容" in part_type or "电容" in pro_desc:
        parameters = extract_parameters("Capacitor", pro_desc)
    elif "电感" in part_type or "电感" in pro_desc:
        parameters = extract_parameters("Inductor", pro_desc)
    elif "轻触开关" in part_type or "轻触开关" in pro_desc:
        parameters = extract_parameters("Tactile Switch", pro_desc)
    elif "LED" in part_type or "LED" in pro_desc:
        parameters = extract_parameters("LED", pro_desc)
    elif "MOSFET" in part_type or "MOSFET" in pro_desc or "场效应管" in part_type or "MOS管" in pro_desc:
        parameters = extract_parameters("MOSFET", pro_desc)
    else:
        pass

    return {"mount_type": mount_type, "part_type": part_type, "specification": parameters['specification'],
            "footprint": footprint, "pro_desc": pro_desc,
            "datasheet_link": datasheet_link}


def query_online(query_file, token, bom_file_path):
    df_temp_bom = pd.read_excel(query_file, skiprows=4, header=None)
    single_query_url = "https://openapi.ickey.cn/search-v1/products/get-single-goods-new"
    query_res = ''
    for index, row in df_temp_bom.iterrows():
        if pd.notna(row[2]):
            pass
        else:
            query_res = single_query(single_query_url, token, row[3])

            if query_res is None:
                # On pending 如果查询不到，那么需要解析原始的
                print("Error: Failed to query.")
                # exit(1)
            else:

                # 对查询结果进行解析
                parser_res = query_res_process(query_res, row[3], row[8])
                # 向当前行的第一列写入parser_res中的字段
                if '贴片' in parser_res['part_type']: parser_res['mount_type'] = ''
                if '陶瓷' in parser_res['part_type']: parser_res['part_type'] = parser_res['part_type'].replace('陶瓷',
                                                                                                                '')
                for item in IC_list:
                    if item in parser_res['part_type']:
                        parser_res['part_type'] = 'IC'
                        break
                # 元件类型填入
                df_temp_bom.at[index, df_temp_bom.columns[2]] = parser_res['mount_type'] + parser_res['part_type']
                # 元件规格填入
                df_temp_bom.at[index, df_temp_bom.columns[3]] = row[3] + ' ' + parser_res['specification'] + ' ' + \
                                                                parser_res['footprint']

                # 查询得到的描述粘贴
                df_temp_bom.at[index, df_temp_bom.columns[12]] = parser_res['pro_desc']
                # 查询得到的规格书链接
                if parser_res['datasheet_link']:
                    df_temp_bom.at[index, df_temp_bom.columns[13]] = 'https:' + parser_res['datasheet_link'][0]

            time.sleep(2)  # 每次查询完成 休眠两秒

    output_file = bom_file_path
    prefix = '自动生成工程BOM'
    name_part, extension = output_file.rsplit('.', 1)
    if extension == 'xls':
        extension = extension.replace('xls', 'xlsx')
    new_filename = name_part + prefix + '.' + extension

    if not os.path.exists(new_filename):
        source_filename = 'source.xlsx'
        source_wb = openpyxl.load_workbook(source_filename)
        source_ws = source_wb.active

        wb = Workbook()
        ws = wb.active

        # 复制前四行数据和样式
        for row in source_ws.iter_rows(min_row=1, max_row=4, max_col=source_ws.max_column):
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = Font(name=cell.font.name,
                                         size=cell.font.size,
                                         bold=cell.font.bold,
                                         italic=cell.font.italic,
                                         vertAlign=cell.font.vertAlign,
                                         underline=cell.font.underline,
                                         strike=cell.font.strike,
                                         color=cell.font.color)
                    new_cell.border = Border(left=cell.border.left,
                                             right=cell.border.right,
                                             top=cell.border.top,
                                             bottom=cell.border.bottom,
                                             diagonal=cell.border.diagonal,
                                             diagonal_direction=cell.border.diagonal_direction,
                                             outline=cell.border.outline,
                                             vertical=cell.border.vertical,
                                             horizontal=cell.border.horizontal)
                    new_cell.fill = PatternFill(fill_type=cell.fill.fill_type,
                                                start_color=cell.fill.start_color,
                                                end_color=cell.fill.end_color)
                    new_cell.number_format = cell.number_format
        wb.save(new_filename)

    refresh_excel(new_filename)
    with pd.ExcelWriter(new_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df_temp_bom.to_excel(writer, index=False, startrow=4, header=False, sheet_name='Sheet')

    color_mark(new_filename)
    # df_temp_bom.to_excel(query_file, startrow=4, index=False, header=False)
