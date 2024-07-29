from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox
from footprint.footprint import initialize_set, add_to_set, remove_from_set, get_footprint_set, any_footprint_in_string
from preProcess.KB_preProcess import KB_manufacturer_process
from preProcess.MC_preProcess import MC_preProcess
from preProcess.mergePreprocess import merge_preprocess
from query import query_online, token_access, refresh_excel
from test import Ui_Dialog
import pandas as pd
from db_model import K3Data
from typing import Dict
import openpyxl
from openpyxl.styles import PatternFill

# 初始化封装集合
initialize_set(
    {'DIP', 'DIL', 'SIP', 'SIL', 'ZIP', 'QFP', 'SOJ', 'PLCC', 'CLCC', 'TSOC', 'PGA', 'BGA', 'SMD', 'SOP', 'SOIC',
     'CERPAC', 'DMP', 'SC70', 'SOT', 'TO', 'QFN', '0201', '0402', '0603', '0805', '1206', '1210', '1812', '2010',
     '2512'})


def remove_last_word(s):
    parts = s.rsplit(' ', 1)
    return parts[0] if len(parts) > 1 else ''


def text_to_val(lineEdit_text) -> int:
    try:
        assert len(lineEdit_text) > 0
        return int(lineEdit_text)
    except AssertionError:
        print("AssertionError : {lineEdit_text}")
    except ValueError:
        print("cannot convert text to number : {lineEdit_text}")


# 通过料号进行查询，返回K3中该料号的相关信息
def query_part_info_by_partcode(part_code) -> Dict[str, str]:
    """
    使用料号在K3数据库中进行查询，返回对应的规格描述
    :param part_code:料号
    :return:该客户、该料号在K3中的规格描述
    """
    query_results = K3Data.select().where(K3Data.specification.contains(part_code))

    if len(query_results) == 0:  # case 1：如果没有查询结果
        return {}
    elif len(query_results) == 1:  # case 2：如果查询结果只有1条，直接返回
        # return {'k3code': query_results[0].k3code, 'type_name': query_results[0].type_name,
        #         'specification': query_results[0].specification}
        return {'k3code': '', 'type_name': query_results[0].type_name,  # 客户不同，不能返回k3code
                'specification': query_results[0].specification}
    else:  # case 3: 如果查询结果有多条，那么从结果集合中返回一条规格最长的
        max_spec_length = 0  # 用于追踪最大长度的变量
        selected_result = None  # 用于存储最大长度对应的结果

        for query_result in query_results:
            current_spec_length = len(query_result.specification)

            if current_spec_length > max_spec_length:
                max_spec_length = current_spec_length
                selected_result = {
                    'k3code': '',  # 客户不匹配，不返回K3 编码
                    'type_name': query_result.type_name,
                    'specification': query_result.specification
                }

        return selected_result


def query_part_info(part_code, customer) -> Dict[str, str]:
    """
    使用料号、客户编码在K3数据库中进行查询，返回对应的规格描述
    :param part_code:料号
    :param customer:客户编码
    :return:该客户、该料号在K3中的规格描述
    """
    # 根据 料号以及客户编码进行查询
    query_results = K3Data.select().where(K3Data.specification.contains(part_code) &
                                          K3Data.specification.contains(customer))

    if len(query_results) == 1:  # case 2：如果查询结果只有1条，直接返回
        return {'k3code': query_results[0].k3code, 'type_name': query_results[0].type_name,
                'specification': query_results[0].specification}
    elif len(query_results) > 1:  # case 3: 如果查询结果有多条，那么从结果集合中返回一条规格最长的
        max_spec_length = 0  # 用于追踪最大长度的变量
        selected_result = None  # 用于存储最大长度对应的结果

        for query_result in query_results:
            current_spec_length = len(query_result.specification)

            if current_spec_length > max_spec_length:
                max_spec_length = current_spec_length
                selected_result = {
                    'k3code': query_result.k3code,
                    'type_name': query_result.type_name,
                    'specification': query_result.specification
                }

        return selected_result
    elif len(query_results) == 0:  # case 1：如果没有查询结果，放宽条件，只用料号来查
        return query_part_info_by_partcode(part_code)
    # print('{} {} on {}'.format(note.id, note.text, note.created))


class MyMainWindow(QMainWindow, Ui_Dialog):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # 原始BOM的元数据信息
        self.customer_code = ""
        self.initial_row_line = -1
        self.end_row_line = -1
        self.part_code_col = -1
        self.manufacturer_col = -1
        self.designator_col = -1
        self.quantity_col = -1
        self.description_col = -1
        self.footprint_col = -1
        self.alternative1_code_col = -1
        self.alternative2_code_col = -1
        self.alternative1_manufacturer_col = -1
        self.alternative2_manufacturer_col = -1

    def on_click(self):  # 按钮被按下
        self.customer_code = self.lineEdit_customer_code.text()  # 客户代码
        if self.lineEdit_initial_row.text() == '':
            self.messageBox_info('起始行为空')
        else:
            self.initial_row_line = text_to_val(self.lineEdit_initial_row.text())  # 开始行

        if self.lineEdit_end_row.text() == '':
            self.messageBox_info('结束行为空')
        else:
            self.end_row_line = text_to_val(self.lineEdit_end_row.text())  # 结束行

        if self.lineEdit_part_code.text() == '':
            self.messageBox_info('料号列为空')
        else:
            self.part_code_col = ord((self.lineEdit_part_code.text())[0]) - ord('A')  # 料号列

        if self.lineEdit_designator.text() == '':
            self.messageBox_info('位号列为空')
        else:
            self.designator_col = ord((self.lineEdit_designator.text())[0]) - ord('A')  # 料号列

        if 'merge' not in self.customer_code:
            if self.lineEdit_quantity.text() == '':
                self.messageBox_info('用量列为空')
            else:
                self.quantity_col = ord((self.lineEdit_quantity.text())[0]) - ord('A')  # 料号列

        if self.lineEdit_description.text() != '':
            self.description_col = ord((self.lineEdit_description.text())[0]) - ord('A')  # 描述列
        if self.lineEdit_manufacturer.text() != '':
            self.manufacturer_col = ord((self.lineEdit_manufacturer.text())[0]) - ord('A')  # 制造商列
        if self.lineEdit_alternative_part_code_1.text() != '':
            self.alternative1_code_col = ord((self.lineEdit_alternative_part_code_1.text())[0]) - ord('A')  # 替代料号1列号
            self.alternative2_code_col = ord((self.lineEdit_alternative_part_code_2.text())[0]) - ord('A')  # 替代料号2列号
            self.alternative1_manufacturer_col = ord((self.lineEdit_alternative_manufacturer_1.text())[0]) - ord(
                'A')  # 替代料号1厂商列号
            self.alternative2_manufacturer_col = ord((self.lineEdit_alternative_manufacturer_2.text())[0]) - ord(
                'A')  # 替代料号2厂商列号

        # 选填字段处理,可填可不填
        if self.lineEdit_footprint.text() != '':
            self.footprint_col = ord((self.lineEdit_footprint.text())[0]) - ord('A')  # 封装列

        self.DB_process()  # K3数据库建立
        self.raw_BOM_copy_process()  # 照抄的列填充
        self.raw_BOM_parse_process()  # 需要查询的列填充
        # 如果存在替代同行的情况，那么执行替代处理
        if self.alternative1_code_col != self.part_code_col:
            refresh_excel('output.xlsx')
            self.alternative_process()

        self.online_query()  # 在线查询

    def messageBox_info(self, info):
        msg_box = QMessageBox()

        msg_box.setText(info)
        msg_box.setWindowTitle("提示")
        msg_box.setStandardButtons(QMessageBox.Ok)
        retval = msg_box.exec()

    def raw_BOM_copy_process(self):
        """
        对原始BOM中不需要解析的字段直接复制粘贴
        :return:
        """
        if self.customer_code == 'MC':  # MC 预处理
            self.BOMInput_textEdit.input_path_file, me_content_len = MC_preProcess(
                self.BOMInput_textEdit.input_path_file)
            self.end_row_line = self.initial_row_line + me_content_len
        if 'merge' in self.customer_code:  # 坐标BOM预处理
            self.BOMInput_textEdit.input_path_file, self.quantity_col = merge_preprocess(
                self.BOMInput_textEdit.input_path_file, self.initial_row_line - 1, self.part_code_col)
        print(self.BOMInput_textEdit.input_path_file)
        excel_file_path = self.BOMInput_textEdit.input_path_file
        df = pd.read_excel(excel_file_path)
        start_row = self.initial_row_line - 2  # 指定起始行索引（从0开始）
        end_row = self.end_row_line - 1
        designator_list = df.iloc[start_row:end_row, self.designator_col].tolist()  # 位号获取
        if self.footprint_col != -1:
            footprint_list = df.iloc[start_row:end_row, self.footprint_col].tolist()  # 封装获取
        print(self.quantity_col)
        quantity_list = df.iloc[start_row:end_row, self.quantity_col].tolist()  # 用量获取

        description_list = df.iloc[start_row:end_row, self.description_col].tolist()  # 描述获取
        manufacturer_list = []
        if self.manufacturer_col != -1:
            manufacturer_list = df.iloc[start_row:end_row, self.manufacturer_col].tolist()  # 制造商获取

        if self.customer_code == 'KB':  # KB预处理
            manufacturer_list = KB_manufacturer_process(description_list)

        refresh_excel('target.xlsx')
        workbook = openpyxl.load_workbook('target.xlsx')

        sheet = workbook.active

        start_row = 5  # 开始写的行

        for i, value in enumerate(designator_list):
            # 位号转换成BQC格式
            if ',' in str(value):
                sheet['H' + str(start_row + i)] = str(value).replace(' ', '').replace(',', ' ')
            else:
                sheet['H' + str(start_row + i)] = str(value)

        # for i, value in enumerate(footprint_list):   #封装暂时不用写
        #     sheet['G' + str(start_row + i)] = value
        for i, value in enumerate(quantity_list):
            sheet['G' + str(start_row + i)] = value

        if self.manufacturer_col != -1:  # 制造商如果原始BOM 没有单独的列，那么可以不写
            for i, value in enumerate(manufacturer_list):
                sheet['I' + str(start_row + i)] = value
        if self.customer_code == 'KB':
            for i, value in enumerate(manufacturer_list):
                sheet['I' + str(start_row + i)] = value
        for i in range(len(designator_list)):
            sheet['A' + str(start_row + i)] = i + 1

        # 保存Excel文件
        workbook.save('target.xlsx')

    # xlsx格式原始BOM处理
    def raw_BOM_parse_process(self):
        """
        对原始BOM中需要解析的字段进行解析处理
        :return:
        """
        print(self.BOMInput_textEdit.input_path_file)
        excel_file_path = self.BOMInput_textEdit.input_path_file
        df = pd.read_excel(excel_file_path)
        column_index = self.part_code_col
        start_row = self.initial_row_line - 2  # 指定起始行索引（从0开始）
        end_row = self.end_row_line - 1
        # column_data = df.iloc[start_row:, column_index].values
        part_code_column = df.iloc[start_row:end_row, column_index].tolist()  # 获取需要查询的料号

        print("查询数量", len(part_code_column))

        ########如果料号和描述混杂成一列，需要通过解析另加功能处理 On pending ############

        for i in range(len(part_code_column)):
            if pd.isnull(part_code_column[i]):  # 防止料号为空
                part_code_column[i] = "料号为空"

            if isinstance(part_code_column[i], (int, float)):  # 将数字类型转换为字符串类型，便于后续的查询
                part_code_column[i] = str(part_code_column[i])

            part_code = part_code_column[i]
            print(i, "  ", part_code)

            query_res = query_part_info(part_code, self.customer_code)
            self.query_res_process(part_code, query_res, i)

    def query_res_process(self, part_code, query_res, part_index):
        """
        :param part_index:
        :param query_res:
        :return:
        """
        # 处理字典类型的查询结果，读取K3代码，元件类型，以及规格描述，从规格描述中按照
        if len(query_res) != 0:
            print(query_res['k3code'], query_res['type_name'], query_res['specification'])
            workbook = openpyxl.load_workbook('target.xlsx')
            sheet = workbook.active
            start_row = 5  # 开始写的行
            sheet['B' + str(start_row + part_index)] = query_res['k3code']

            sheet['C' + str(start_row + part_index)] = query_res['type_name']
            temp_customer_code = query_res['specification'].split()[0]
            # sheet['C' + str(start_row + part_index)] = query_res['specification'].replace(self.customer_code + " ", "")
            sheet['D' + str(start_row + part_index)] = query_res['specification'].replace(temp_customer_code + " ",
                                                                                          "")  # 从k3 规格中去掉客户编码
            workbook.save('target.xlsx')

        else:  # 无查询结果，在规格栏写上料号
            workbook = openpyxl.load_workbook('target.xlsx')
            sheet = workbook.active
            start_row = 5  # 开始写的行
            sheet['D' + str(start_row + part_index)] = part_code
            workbook.save('target.xlsx')

    def DB_process(self):
        """
        读取用户输入的K3文件，建立查询的数据库
        :return:
        """
        K3Data.truncate_table()
        print(K3Data.select().count())
        excel_file_path = self.K3Input_textEdit.input_path_file
        df = pd.read_excel(excel_file_path)

        data = [{'k3code': row['K3Code'], 'type_name': row['Name'], 'specification': row['Specification']} for _, row in
                df.iterrows()]
        K3Data.bulk_create([K3Data(**row) for row in data], batch_size=100)  # 每批次插入100条数据
        # for index, row in df.iterrows():
        #   note = K3Data.create(k3code=row['K3Code'], type_name=row['Name'], specification=row['Specification'])
        # 将读入的K3BOM建立查询的数据库
        # K3Data.insert(df).execute()
        print(K3Data.select().count())

    def alternative_process(self):
        excel_file_path = self.BOMInput_textEdit.input_path_file
        df_initial_bom_path = pd.read_excel(excel_file_path)  # 读取原始BOM
        df_res_bom = pd.read_excel('target.xlsx', skiprows=4, header=None)  # 读取工程BOM

        # 从原始BOM读取替代信息
        start_row = self.initial_row_line - 2  # 指定起始行索引（从0开始）
        end_row = self.end_row_line - 1
        alter1_code_list = df_initial_bom_path.iloc[start_row:end_row, self.alternative1_code_col].tolist()  # 替代1料号
        manu1_list = df_initial_bom_path.iloc[start_row:end_row, self.alternative1_manufacturer_col].tolist()  # 替代1厂商
        alter2_code_list = df_initial_bom_path.iloc[start_row:end_row, self.alternative2_code_col].tolist()  # 替代2料号
        manu2_list = df_initial_bom_path.iloc[start_row:end_row, self.alternative2_manufacturer_col].tolist()  # 替代2厂商

        res_bom_index = 0
        # 遍历替代，如果有替代，则往工程BOM中进行插入操作
        for index, code in enumerate(alter1_code_list):
            if not pd.isnull(code):
                row_to_copy = df_res_bom.iloc[res_bom_index:res_bom_index + 1]  # 获取当前行数据
                df_res_bom = pd.concat([df_res_bom.iloc[:res_bom_index + 1], row_to_copy,
                                        df_res_bom.iloc[res_bom_index + 1:]]).reset_index(drop=True)  # 在其下复制一行
                res_bom_index = res_bom_index + 1

                df_res_bom.iloc[res_bom_index, 8] = manu1_list[index]  # 替代1厂商填写

                df_res_bom.iloc[res_bom_index, 1] = ''

                query_res_alter1 = query_part_info(str(code), self.customer_code)  # 替代1 K3 规格查询
                if len(query_res_alter1) != 0:
                    df_res_bom.iloc[res_bom_index, 1] = query_res_alter1['k3code']  # 替代1 k3code填写
                    df_res_bom.iloc[res_bom_index, 2] = query_res_alter1['type_name']  # 替代1 类型填写
                    df_res_bom.iloc[res_bom_index, 3] = query_res_alter1['specification'].replace(
                        self.customer_code + " ", "")  # 替代1 规格填写

                    # 检查主料有无查询结果，如果无，那么用替代1 信息复制过去
                    if pd.isnull(df_res_bom.iloc[res_bom_index - 1, 2]):
                        df_res_bom.iloc[res_bom_index - 1, 2] = query_res_alter1['type_name']
                        temp_part_code = df_res_bom.iloc[res_bom_index - 1, 3]  # 主料料号
                        temp_manu = df_res_bom.iloc[res_bom_index - 1, 8]  # 主料供应商
                        last_word_is_footprint = any_footprint_in_string(
                            query_res_alter1['specification'].split()[-1])  # 规格的最后一个单词 是否是封装

                        if last_word_is_footprint:  # 如果最后一个词是封装，直接复制过去即可
                            df_res_bom.iloc[res_bom_index - 1, 3] = query_res_alter1['specification'].replace(
                                self.customer_code + " ", "").replace(str(code), temp_part_code)+' '+temp_manu
                        else:  # 如果最后一个单词是品牌
                            df_res_bom.iloc[res_bom_index - 1, 3] = remove_last_word(
                                query_res_alter1['specification']).replace(
                                self.customer_code + " ", "").replace(str(code), temp_part_code) + ' ' + temp_manu
                        # On pending 复制 需要去掉原来的 供应商

                else:  # 替代一查询无果，查询主料号记录，复制替代1里
                    if not pd.isnull(df_res_bom.iloc[res_bom_index - 1, 2]):
                        df_res_bom.iloc[res_bom_index, 2] = df_res_bom.iloc[res_bom_index - 1, 2]  # 使用主料号的类型
                        temp_part_code = df_res_bom.iloc[res_bom_index - 1, 3].split()[0]  # 主料的料号
                        temp_manu = df_res_bom.iloc[res_bom_index, 8]  # 替代1 的品牌
                        last_word_is_footprint = any_footprint_in_string(
                            df_res_bom.iloc[res_bom_index - 1, 3].split()[-1])  # 规格的最后一个单词 是否是封装
                        if last_word_is_footprint:
                            df_res_bom.iloc[res_bom_index, 3] = df_res_bom.iloc[res_bom_index - 1, 3].replace(
                                temp_part_code, str(code))+' '+temp_manu  # 替代1 规格填写
                        else:
                            df_res_bom.iloc[res_bom_index, 3] = remove_last_word(df_res_bom.iloc[res_bom_index - 1, 3]).replace(
                                temp_part_code, str(code))+' '+temp_manu  # 替代1 规格填写
                        # On pending 复制 需要去掉原来的 供应商

                    else:
                        # 主料号也无查询结果，暂时填一个替代一的物料号
                        df_res_bom.iloc[res_bom_index, 3] = str(code)

                if not pd.isnull(alter2_code_list[index]):  # 替代二也有东西，再复制一行
                    df_res_bom = pd.concat([df_res_bom.iloc[:res_bom_index + 1], row_to_copy,
                                            df_res_bom.iloc[res_bom_index + 1:]]).reset_index(drop=True)  # 在其下复制一行
                    res_bom_index = res_bom_index + 1

                    df_res_bom.iloc[res_bom_index, 8] = manu2_list[index]  # 替代2厂商填写
                    df_res_bom.iloc[res_bom_index, 1] = ''

                    query_res_alter2 = query_part_info(str(alter2_code_list[index]), self.customer_code)  # 替代2 K3 规格查询
                    if len(query_res_alter2) != 0:
                        df_res_bom.iloc[res_bom_index, 1] = query_res_alter2['k3code']  # 替代2 类型填写
                        df_res_bom.iloc[res_bom_index, 2] = query_res_alter2['type_name']  # 替代2 类型填写
                        df_res_bom.iloc[res_bom_index, 3] = query_res_alter2['specification'].replace(
                            self.customer_code + " ", "")  # 替代2 规格填写
                        # 检查主料、替代1 有无查询结果，如果无，那么填补

                        # 检查替代一有无查询结果，如果无，那么用替代二填补
                        if pd.isnull(df_res_bom.iloc[res_bom_index - 1, 2]):
                            df_res_bom.iloc[res_bom_index - 1, 2] = query_res_alter2['type_name']
                            temp_part_code = df_res_bom.iloc[res_bom_index - 1, 3]
                            temp_manu = df_res_bom.iloc[res_bom_index - 1, 8]
                            last_word_is_footprint = any_footprint_in_string(query_res_alter2['specification'].split()[-1])
                            if last_word_is_footprint:
                                df_res_bom.iloc[res_bom_index - 1, 3] = query_res_alter2['specification'].replace(
                                    self.customer_code + " ", "").replace(str(alter2_code_list[index]), temp_part_code) +' '+temp_manu
                            else:
                                df_res_bom.iloc[res_bom_index - 1, 3] = remove_last_word(query_res_alter2['specification']).replace(
                                    self.customer_code + " ", "").replace(str(alter2_code_list[index]), temp_part_code) +' '+temp_manu
                        # 检查主料有无查询结果，如果无，那么用替代填补
                        if pd.isnull(df_res_bom.iloc[res_bom_index - 2, 2]):
                            df_res_bom.iloc[res_bom_index - 2, 2] = query_res_alter2['type_name']
                            temp_part_code = df_res_bom.iloc[res_bom_index - 2, 3]
                            temp_manu = df_res_bom.iloc[res_bom_index - 2, 8]
                            last_word_is_footprint = any_footprint_in_string(query_res_alter2['specification'].split()[-1])
                            if last_word_is_footprint:
                                df_res_bom.iloc[res_bom_index - 2, 3] = query_res_alter2['specification'].replace(
                                    self.customer_code + " ", "").replace(str(alter2_code_list[index]), temp_part_code)+' '+temp_manu
                            else:
                                df_res_bom.iloc[res_bom_index - 2, 3] = remove_last_word(query_res_alter2['specification']).replace(
                                    self.customer_code + " ", "").replace(str(alter2_code_list[index]), temp_part_code)+' '+temp_manu
                    else:
                        if not pd.isnull(df_res_bom.iloc[res_bom_index - 1, 2]):
                            df_res_bom.iloc[res_bom_index, 2] = df_res_bom.iloc[res_bom_index - 1, 2]  # 替代2 类型填写
                            temp_part_code = df_res_bom.iloc[res_bom_index - 1, 3].split()[0]
                            temp_manu = df_res_bom.iloc[res_bom_index, 8]
                            last_word_is_footprint = any_footprint_in_string(df_res_bom.iloc[res_bom_index - 1, 3].split()[-1])
                            if last_word_is_footprint:
                                df_res_bom.iloc[res_bom_index, 3] = df_res_bom.iloc[res_bom_index - 1, 3].replace(
                                    temp_part_code, str(alter2_code_list[index]))+' '+ temp_manu  # 替代2 规格填写
                            else:
                                df_res_bom.iloc[res_bom_index, 3] = remove_last_word(df_res_bom.iloc[res_bom_index - 1, 3]).replace(
                                    temp_part_code, str(alter2_code_list[index]))+' '+ temp_manu  # 替代2 规格填写
                        else:
                            # 主料号也无查询结果，暂时填一个替代二的物料号
                            df_res_bom.iloc[res_bom_index, 3] = str(alter2_code_list[index])
                res_bom_index = res_bom_index + 1

            else:
                res_bom_index = res_bom_index + 1

        with pd.ExcelWriter('output.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_res_bom.to_excel(writer, index=False, startrow=4, header=False)

    def online_query(self):
        token = token_access()
        if self.alternative1_code_col == -1:
            query_online('target.xlsx', token, self.BOMInput_textEdit.input_path_file)
        else:
            query_online('output.xlsx', token, self.BOMInput_textEdit.input_path_file)
        print("process complete")


if __name__ == "__main__":
    app = QApplication([])
    window = MyMainWindow()
    window.show()
    app.exec()
